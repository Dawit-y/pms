"""
Excel and PDF exporter sanity checks.
"""

from openpyxl import load_workbook

from pms_api.reports.base import ColumnSpec
from pms_api.reports.base import ReportDefinition
from pms_api.reports.exporters import ExcelExporter
from pms_api.reports.exporters import PdfExporter


class _FixtureReport(ReportDefinition):
    code = "_test_fixture"
    name_en = "Fixture Report"
    columns = [
        ColumnSpec("name", "Name", "ስም", "Maqaa"),
        ColumnSpec("count", "Count", "ብዛት", "Lakkoofsa", pdf_align="RIGHT"),
    ]

    def build_queryset(self, params, user):
        return []


def _rows():
    return [
        {"name": "Alpha", "count": 1},
        {"name": "Beta", "count": 2},
        {"name": "Gamma", "count": 3},
    ]


def test_excel_exporter_produces_loadable_workbook(tmp_path):
    exporter = ExcelExporter(_FixtureReport())
    exporter.begin()
    exporter.write_chunk(_rows())
    exporter.write_summary({"total": 6})
    path, size = exporter.finalize()

    assert size > 0
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # rows[0] = title, rows[1] = header, rows[2..4] = data, then blank + summary
    assert rows[0][0].startswith("Fixture Report")
    assert "Name" in rows[1][0]
    data_values = [r[0] for r in rows[2:5]]
    assert data_values == ["Alpha", "Beta", "Gamma"]


def test_pdf_exporter_produces_pdf_bytes(tmp_path):
    exporter = PdfExporter(_FixtureReport())
    exporter.begin()
    exporter.write_chunk(_rows())
    exporter.write_summary({"total": 6})
    path, size = exporter.finalize()

    assert size > 0
    with path.open("rb") as f:
        head = f.read(4)
    assert head == b"%PDF"
