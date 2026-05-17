"""
Excel exporter — openpyxl Workbook in write-only mode.

write_only=True streams rows straight to the .xlsx zip without keeping every
cell in memory, which is the only way to safely export hundreds of thousands
of rows on a worker without OOMing.
"""

import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .base import Exporter


class ExcelExporter(Exporter):
    extension = "xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def __init__(self, definition) -> None:
        super().__init__(definition)
        self._wb: Workbook | None = None
        self._ws = None
        self._has_summary = False

    def begin(self) -> None:
        self._wb = Workbook(write_only=True)
        ws = self._wb.create_sheet(title=(self.definition.name_en or "Report")[:31])
        # _ws must be set BEFORE creating any WriteOnlyCell — the style
        # descriptors walk cell.parent.parent during attribute assignment.
        self._ws = ws

        # Column widths (openpyxl write_only requires column_dimensions before rows)
        for idx, col in enumerate(self.definition.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = col.width

        # Title row spans the whole table.
        title_cell = self._make_cell(
            self.definition.name_en or self.definition.code,
            font=Font(name="Calibri", size=14, bold=True),
        )
        ws.append([title_cell] + [None] * (len(self.definition.columns) - 1))

        # Tri-lingual header row (en / am / or stacked) so reviewers don't have
        # to flip locales.
        header_cells = []
        for col in self.definition.columns:
            label_parts = [p for p in (col.label_en, col.label_am, col.label_or) if p]
            label = "\n".join(label_parts)
            header_cells.append(
                self._make_cell(label, fill=self.HEADER_FILL, font=self.HEADER_FONT),
            )
        ws.append(header_cells)
        ws.row_dimensions[2].height = 35

    def _make_cell(self, value, font=None, fill=None, align=None):
        from openpyxl.cell import WriteOnlyCell  # noqa: PLC0415

        cell = WriteOnlyCell(self._ws, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        cell.alignment = align or self.HEADER_ALIGN
        return cell

    def write_chunk(self, rows: list[dict]) -> None:
        if self._ws is None:
            msg = "ExcelExporter.write_chunk called before begin()."
            raise RuntimeError(msg)
        for row in rows:
            self._ws.append(
                [self._format_cell(col, row.get(col.key)) for col in self.definition.columns],
            )

    @staticmethod
    def _format_cell(col, value):
        if value is None:
            return ""
        if col.formatter is not None:
            try:
                return col.formatter(value)
            except (TypeError, ValueError):
                return value
        return value

    def write_summary(self, summary: dict) -> None:
        if not summary or self._ws is None:
            return
        self._has_summary = True
        # blank separator row
        self._ws.append([None] * len(self.definition.columns))
        header = self._make_cell(
            "Summary",
            font=Font(name="Calibri", size=12, bold=True),
        )
        self._ws.append([header] + [None] * (len(self.definition.columns) - 1))
        for key, value in summary.items():
            label = key.replace("_", " ").title()
            self._ws.append(
                [label, str(value)] + [None] * max(0, len(self.definition.columns) - 2),
            )

    def finalize(self) -> tuple[Path, int]:
        if self._wb is None:
            msg = "ExcelExporter.finalize called before begin()."
            raise RuntimeError(msg)
        with tempfile.NamedTemporaryFile(
            prefix="report_",
            suffix=".xlsx",
            delete=False,
        ) as tmp:
            self._path = Path(tmp.name)
        self._wb.save(self._path)
        size = self._path.stat().st_size
        return self._path, size
